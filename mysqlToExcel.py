#모듈을 불러온다
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.compat import range
import pymysql

#워크북을 하나 만든다
wb = Workbook()

#활성화된 엑셀 시트를 선택한다.
ws = wb.active

#엑셀 시트 제목을 지정한다.
ws.title = '20180424'

#DB 연결 계정
server = '172.16.***.***'
user = 'sgedev***'
password = '*****'
dbname = 'ish****'

#DB 연결
conn = pymysql.connect(server, user, password, dbname, charset='utf8')

#커서를 만든다.
cursor = conn.cursor()

ws['A1'] = "주문번호"
ws['B1'] = "주문일자"
ws['C1'] = "학부모아이디"
ws['D1'] = "학생아이디"
ws['E1'] = "학생로그인아이디"
ws['F1'] = "학습시작일"
ws['G1'] = "학습종료일"
ws['H1'] = "주문상태코드"
ws['I1'] = "주문상태"
ws['J1'] = "주문등록일"

#테이블 내용 저장
cursor.execute("SELECT t1.purchase_no ,t1.purchase_dt ,t1.purchase_id ,t1.student_id ,s.login_id ,t1.start_de ,t1.end_de ,t1.status_gbn ,t3.code_nm ,t1.reg_dt FROM tz_purchase t1 INNER JOIN tz_product t2 ON t1.product_no=t2.product_no AND t2.study_type=40 INNER JOIN tz_code t3 ON t1.status_gbn=t3.code_no AND t3.code_gbn=10118 LEFT OUTER JOIN tz_member s ON t1.student_id = s.user_id where t1.purchase_dt < '2018-01-01 00:00:00' and t1.start_de is null")

#2번째 행을 나타낸다.
row_num = 2

#한행씩 가져온다.
row = cursor.fetchone()
while row:
    column_char = 'a'

    #1~6까지 x가 변하면서 컬럼 문자, row를 하나씩 늘여 결과를 하나씩 담는다.
    #ws['a1'] = row[0], ws['b1'] = row[1]........
    for x in range(1, 11):
        ws[column_char + str(row_num)] = row[x-1]
        column_char = chr(ord(column_char) + 1)

    #다음 행을 표시하기 위해 뒤의 숫자를 증가 시킨다.
    row_num = row_num + 1
    row = cursor.fetchone()

#파일을 실제 저장한다.
wb.save("result_20180424_id.xlsx")

print("엑셀저장 완료")

